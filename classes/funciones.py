import requests
import mysql.connector
from passwords import *
import os
import pandas as pd

#--------------------------------------------------------------------------- Funciones que generan datos para app web

def airtable_url(table_name):
    base_id=joints_base_id
    
    url=f'https://api.airtable.com/v0/{base_id}/{table_name}'
    return url

def configure_headers():
    api_key=adri_jointsplayground_api_key
    
    headers={'Authorization': f'Bearer {api_key}'}
    
    return headers

def rt_connection():
    config = {
        'user': user_rt,
        'password': password_rt,
        'host': host_rt,
        'database': database_components,
        'port': port,        
    }
    connection = mysql.connector.connect(**config)
    #print (config)
    return connection
    
def get_joints():
    
    url=airtable_url(joints_table)
    headers=configure_headers()
    
    # Variables para controlar la paginación y contar registros
    params = {
        'pageSize': 100,  # Tamaño máximo de una página (ajústalo según tus necesidades)
    }
    
    records_list = []  # Lista para almacenar los datos que se devolverán como respuesta

    try:
        offset = None
        while True:
            # Configurar el offset para obtener la siguiente página
            if offset:
                params['offset'] = offset

            # Realizar una petición GET a la tabla de Airtable
            response = requests.get(url, headers=headers, params=params)

            # Verificar si la petición fue exitosa (código de estado 200)
            if response.status_code == 200:
                # Obtener los datos de la página actual en formato JSON
                data = response.json()

                # Acceder a los registros de la página actual
                records = data['records']

                # Si no hay más registros, salir del bucle
                if not records:
                    break

                # Procesar los registros que cumplen la condición y agregarlos a records_list
                for record in records:
                    # Cada registro tiene un campo 'fields' con los datos reales
                    fields = record['fields']

                    # Obtener el valor de la columna 'joint_type_id'
                    joint_type_id = fields.get('joint_type_id')
                    joints3_type = fields.get('JointType_api')
                    cgt = fields.get('api_ConnectionGroup_type')
                    components = fields.get('Components_api')
                    Project = fields.get('project_export')
                    

                    # Agregar el valor a records_list
                    records_list.append({
                        'joint_type_id': joint_type_id,
                        'joints3_type': joints3_type,
                        'api_ConnectionGroup_type': cgt,
                        'Components_api': components,
                        'project_export': Project,                        
                    })
                    
                    #print(records_list)

                # Obtener el offset para la siguiente página, si no hay más, el valor será None y salimos del bucle
                offset = data.get('offset')
                if not offset:
                    break
            else:
                # Si la petición no fue exitosa, imprimir el mensaje de error
                records_list = [f"Error al obtener datos: {response.status_code} - {response.text}"]
                break

        # Ordenar registros
        records_list = sorted(records_list, key=lambda x: x['joint_type_id'])
        

    except requests.exceptions.RequestException as e:
        # Manejar excepciones de conexión
        records_list = [f"Error de conexión: {e}"]

    return records_list

def get_joint_layers(joint):
    
    url=airtable_url(jointlayers_table)
    headers=configure_headers()
    
    # Variables para controlar la paginación y contar registros
    params = {
        'filterByFormula': "SEARCH('"+joint+"', {joint_type_code})",
        'pageSize': 100,  # Tamaño máximo de una página (ajústalo según tus necesidades)
    }
    
    records_list = []  # Lista para almacenar los datos que se devolverán como respuesta

    try:
        offset = None
        while True:
            # Configurar el offset para obtener la siguiente página
            if offset:
                params['offset'] = offset

            # Realizar una petición GET a la tabla de Airtable
            response = requests.get(url, headers=headers, params=params)

            # Verificar si la petición fue exitosa (código de estado 200)
            if response.status_code == 200:
                # Obtener los datos de la página actual en formato JSON
                data = response.json()

                # Acceder a los registros de la página actual
                records = data['records']

                # Si no hay más registros, salir del bucle
                if not records:
                    break

                # Procesar los registros que cumplen la condición y agregarlos a records_list
                for record in records:
                    # Cada registro tiene un campo 'fields' con los datos reales
                    fields = record['fields']

                    # Obtener el valor de las columnas
                    material_id = fields.get('material_id')
                    description = fields.get('Long_Name (from Material)')
                    formula = fields.get('Calculation Formula')
                    performance = fields.get('Performance')

                    # Agregar el valor a records_list
                    records_list.append({
                        'material_id': material_id,
                        'Long_Name': description,
                        'formula': formula,
                        'performance': performance,                                               
                    })
                    
                    #print(records_list)

                # Obtener el offset para la siguiente página, si no hay más, el valor será None y salimos del bucle
                offset = data.get('offset')
                if not offset:
                    break
            else:
                # Si la petición no fue exitosa, imprimir el mensaje de error
                records_list = [f"Error al obtener datos: {response.status_code} - {response.text}"]
                break

    except requests.exceptions.RequestException as e:
        # Manejar excepciones de conexión
        records_list = [f"Error de conexión: {e}"]

    return records_list    
    
def get_connectiontype(cgt):
    
    url=airtable_url(connectiontype_table)
    headers=configure_headers()
    
    # Variables para controlar la paginación y contar registros
    params = {
        'filterByFormula': "SEARCH('"+cgt+"', {connectiongroup_type_id})",
        'pageSize': 100,  # Tamaño máximo de una página (ajústalo según tus necesidades)
    }
    
    records_list = []  # Lista para almacenar los datos que se devolverán como respuesta

    try:
        offset = None
        while True:
            # Configurar el offset para obtener la siguiente página
            if offset:
                params['offset'] = offset

            # Realizar una petición GET a la tabla de Airtable
            response = requests.get(url, headers=headers, params=params)

            # Verificar si la petición fue exitosa (código de estado 200)
            if response.status_code == 200:
                # Obtener los datos de la página actual en formato JSON
                data = response.json()

                # Acceder a los registros de la página actual
                records = data['records']

                # Si no hay más registros, salir del bucle
                if not records:
                    break
                
                # Procesar los registros que cumplen la condición y agregarlos a records_list
                for record in records:
                    # Cada registro tiene un campo 'fields' con los datos reales
                    fields = record['fields']

                    # Obtener el valor de las columnas
                    connection_type_id = fields.get('connection_type')
                    description = fields.get('description (from connection_type_id)')
                    formula = fields.get('Calculation Formula')
                    performance = fields.get('Performance')
                    cgtdesc = fields.get('CGT_description')

                    # Agregar el valor a records_list
                    records_list.append({
                        'connection_type_id': connection_type_id,
                        'description': description,
                        'formula': formula,
                        'performance': performance,
                        'CGtype_description':cgtdesc                                               
                    })
                    
                    
                # Obtener el offset para la siguiente página, si no hay más, el valor será None y salimos del bucle
                offset = data.get('offset')
                if not offset:
                    break
            else:
                # Si la petición no fue exitosa, imprimir el mensaje de error
                records_list = [f"Error al obtener datos: {response.status_code} - {response.text}"]
                break

    except requests.exceptions.RequestException as e:
        # Manejar excepciones de conexión
        records_list = [f"Error de conexión: {e}"]

    return records_list

def getcomp():
    connection=rt_connection()
    cursor= connection.cursor()
    
    query = "select code, status, description from component_type where status <> 'deprecated' order by code"

    cursor.execute(query)
    
    results = cursor.fetchall()
    
    cursor.close()
    connection.close()
    
    return results

def getExecutionUnits(component_type):
    connection=rt_connection()
    cursor= connection.cursor()
    
    query = "select execution_unit_type.code, execution_unit_type.description, execution_unit_type.thickness "\
      "from execution_unit_type "\
      "inner join component_type_execution_unit_type on execution_unit_type.id=component_type_execution_unit_type.execution_unit_type_id "\
      "inner join component_type on component_type_execution_unit_type.component_type_id=component_type.id "\
      f"where component_type.code='{component_type}'"
      
    cursor.execute(query)
    
    results = cursor.fetchall()
    
    cursor.close()
    connection.close()
    
    return results

def getLayerGroups(EU):
    connection=rt_connection()
    cursor= connection.cursor()
    
    query = "select layer_group_type.code, layer_group_type.description "\
      "from execution_unit_type "\
      "inner join execution_unit_type_layer_group_type on execution_unit_type.id=execution_unit_type_layer_group_type.execution_unit_type_id "\
      "inner join layer_group_type on execution_unit_type_layer_group_type.layer_group_type_id=layer_group_type.id "\
      f"where execution_unit_type.code='{EU}'"
      
    cursor.execute(query)
    
    results = cursor.fetchall()
    
    cursor.close()
    connection.close()
    
    return results

def getcgclass():
    
    url=airtable_url(cgclass_table)
    headers=configure_headers()
    
    # Variables para controlar la paginación y contar registros
    params = {
        'pageSize': 100,  # Tamaño máximo de una página (ajústalo según tus necesidades)
    }
    
    records_list = []  # Lista para almacenar los datos que se devolverán como respuesta

    try:
        offset = None
        while True:
            # Configurar el offset para obtener la siguiente página
            if offset:
                params['offset'] = offset

            # Realizar una petición GET a la tabla de Airtable
            response = requests.get(url, headers=headers, params=params)

            # Verificar si la petición fue exitosa (código de estado 200)
            if response.status_code == 200:
                # Obtener los datos de la página actual en formato JSON
                data = response.json()

                # Acceder a los registros de la página actual
                records = data['records']

                # Si no hay más registros, salir del bucle
                if not records:
                    break

                # Procesar los registros que cumplen la condición y agregarlos a records_list
                for record in records:
                    # Cada registro tiene un campo 'fields' con los datos reales
                    fields = record['fields']

                    # Obtener el valor de la columna 'joint_type_id'
                    connectiongroup_class = fields.get('connectiongroup_class')
                    Description_CGClass = fields.get('Description_CGClass')
                    planilla = fields.get('Planilla PDF')
                    box_type = fields.get('box_type (from boxtype_id)')
                                 
                    # Agregar el valor a records_list
                    records_list.append({
                        'connectiongroup_class': connectiongroup_class,
                        'Description_CGClass': Description_CGClass,
                        'planilla': planilla,  
                        'box_type': box_type,                                           
                    })
                    
                    #print(records_list)

                # Obtener el offset para la siguiente página, si no hay más, el valor será None y salimos del bucle
                offset = data.get('offset')
                if not offset:
                    break
            else:
                # Si la petición no fue exitosa, imprimir el mensaje de error
                records_list = [f"Error al obtener datos: {response.status_code} - {response.text}"]
                break

        # Ordenar registros
        records_list = sorted(records_list, key=lambda x: x['box_type'], reverse=True)
        

    except requests.exceptions.RequestException as e:
        # Manejar excepciones de conexión
        records_list = [f"Error de conexión: {e}"]

    return records_list

def getcgtype(cgclass):
    url=airtable_url(cgtype_table)
    headers=configure_headers()
    
    params = {
        'filterByFormula': "SEARCH('"+cgclass+"', {api_ConnectionGroup_Class})",
        'pageSize': 100,  # Tamaño máximo de una página (ajústalo según tus necesidades)
    }
    
    records_list = []  # Lista para almacenar los datos que se devolverán como respuesta

    try:
        offset = None
        while True:
            # Configurar el offset para obtener la siguiente página
            if offset:
                params['offset'] = offset

            # Realizar una petición GET a la tabla de Airtable
            response = requests.get(url, headers=headers, params=params)

            # Verificar si la petición fue exitosa (código de estado 200)
            if response.status_code == 200:
                # Obtener los datos de la página actual en formato JSON
                data = response.json()

                # Acceder a los registros de la página actual
                records = data['records']

                # Si no hay más registros, salir del bucle
                if not records:
                    break

                # Procesar los registros que cumplen la condición y agregarlos a records_list
                for record in records:
                    # Cada registro tiene un campo 'fields' con los datos reales
                    fields = record['fields']

                    # Obtener el valor de la columna 'joint_type_id'
                    cgtype_id = fields.get('cgtype_id')
                    connectiongroup_class = fields.get('api_ConnectionGroup_Class')
                    cgclass_description = fields.get('CG_Class_api')
                    description = fields.get('Description')
                    connectiontype = fields.get('RL_cgtype_ctype (from ConnectionGroup_type)')
                                 
                    # Agregar el valor a records_list
                    records_list.append({
                        'cgtype_id': cgtype_id,
                        'connectiongroup_class': connectiongroup_class,
                        'cgclass_description': cgclass_description,  
                        'description': description,  
                        'connectiontype': connectiontype                                         
                    })
                                        
                # Obtener el offset para la siguiente página, si no hay más, el valor será None y salimos del bucle
                offset = data.get('offset')
                if not offset:
                    break
            else:
                # Si la petición no fue exitosa, imprimir el mensaje de error
                records_list = [f"Error al obtener datos: {response.status_code} - {response.text}"]
                break

        # Ordenar registros
        records_list = sorted(records_list, key=lambda x: x['cgtype_id'])
        

    except requests.exceptions.RequestException as e:
        # Manejar excepciones de conexión
        records_list = [f"Error de conexión: {e}"]

    return records_list

def get_clayers(cgt):
    ct_info=get_connectiontype(cgt)
    
    ct=[]
    
    for i in ct_info:
        ct.append(i['connection_type_id'])
    
    url=airtable_url(clayers_table)
    headers=configure_headers()
    
    records_list = []  # Lista para almacenar los datos que se devolverán como respuesta
    
    for i in ct:
        params = {
            'filterByFormula': "SEARCH('"+i+"', {connection_type_code})",
            'pageSize': 100,  # Tamaño máximo de una página (ajústalo según tus necesidades)
        }      

        try:
            offset = None
            while True:
                # Configurar el offset para obtener la siguiente página
                if offset:
                    params['offset'] = offset

                # Realizar una petición GET a la tabla de Airtable
                response = requests.get(url, headers=headers, params=params)

                # Verificar si la petición fue exitosa (código de estado 200)
                if response.status_code == 200:
                    # Obtener los datos de la página actual en formato JSON
                    data = response.json()

                    # Acceder a los registros de la página actual
                    records = data['records']

                    # Si no hay más registros, salir del bucle
                    if not records:
                        break

                    # Procesar los registros que cumplen la condición y agregarlos a records_list
                    for record in records:
                        # Cada registro tiene un campo 'fields' con los datos reales
                        fields = record['fields']

                        # Obtener el valor de la columna 'joint_type_id'
                        connectiontype_id = fields.get('connection_type_code')
                        material = fields.get('material_id')
                        description = fields.get('Long name (from Material)')
                        performance = fields.get('Performance')
                        formula = fields.get('Calculation Formula')
                        fase=fields.get('Fase')
                                    
                        # Agregar el valor a records_list
                        records_list.append({
                            'connectiontype_id': connectiontype_id,
                            'material': material,
                            'description': description, 
                            'performance': performance,  
                            'formula': formula,  
                            'fase': fase                                         
                        })
                                            
                    # Obtener el offset para la siguiente página, si no hay más, el valor será None y salimos del bucle
                    offset = data.get('offset')
                    if not offset:
                        break
                else:
                    # Si la petición no fue exitosa, imprimir el mensaje de error
                    records_list = [f"Error al obtener datos: {response.status_code} - {response.text}"]
                    break      
            
        except requests.exceptions.RequestException as e:
            # Manejar excepciones de conexión
            records_list = [f"Error de conexión: {e}"]

    # Ordenar registros
    records_list = sorted(records_list, key=lambda x: x['connectiontype_id'])

    #return records_list
    return records_list


#--------------------------------------------------------------------------- LECTURA IFC cajas y conexiones

import ifcopenshell
import ifcopenshell.util
import ifcopenshell.util.element
import openpyxl, math

def ruta_corrector(ruta):
    ruta_corregida=ruta.replace('\\','/')
    return ruta_corregida

def abrir_ifc(ruta):
    ifc=ifcopenshell.open(ruta)
    return ifc

def list_elements(ifc, type, pset, parameter, value):
    boxes=[]
    building_elem_proxy_collector = ifc.by_type(type)
    for i in building_elem_proxy_collector:
        i_psets=ifcopenshell.util.element.get_psets(i)
        if pset in i_psets:
            parameter_value=i_psets[pset].get(parameter, '')
            if value in parameter_value:
                boxes.append(i)
    return boxes


def boxes_info_joint(ruta):
    
    boxes_info=[]
    
    ifc=abrir_ifc(ruta)
    type='IfcBuildingElementProxy'
    pset='EI_Elements Identification'
    parameter='EI_Type'
    value='Joint'
    boxes=list_elements(ifc, type, pset, parameter, value)
    
    
    for i in boxes:
        i_psets = ifcopenshell.util.element.get_psets(i)
        JS_Joint_Specification = i_psets.get('JS_Joint Specification', {})
        QU_Quantity=i_psets.get('QU_Quantity',{})
        Pset_QuantityTakeOff=i_psets.get('Pset_QuantityTakeOff',{})
        JointTypeID = JS_Joint_Specification.get('JS_JointTypeID', '')
        cgt=JS_Joint_Specification.get('JS_ConnectionGroupTypeID', '')
        parent_joint_id = JS_Joint_Specification.get('JS_ParentJointInstanceID', '')
        r_guid = i_psets['EI_Interoperability'].get('RevitGUID', '')
        QU_Length=QU_Quantity.get('QU_Length_m', 0)
        Box_type=Pset_QuantityTakeOff.get('Reference', '')
        inst_a=JS_Joint_Specification.get('JS_C01_ID', '')
        inst_b=JS_Joint_Specification.get('JS_C02_ID', '')
        corematgroup=JS_Joint_Specification.get('JS_CORE_matgroup','')
        Q1matgroup=JS_Joint_Specification.get('JS_Q1_matgroup','')
        Q2matgroup=JS_Joint_Specification.get('JS_Q2_matgroup','')
        Q3matgroup=JS_Joint_Specification.get('JS_Q3_matgroup','')
        Q4matgroup=JS_Joint_Specification.get('JS_Q4_matgroup','')
        parameters_info={'RevitGUID':r_guid, 'JS_ParentJointInstanceID':parent_joint_id, 'JS_JointTypeID':JointTypeID, 'JS_ConnectionGroupTypeID':cgt, 'Core Matgroup':corematgroup,'Q1 Matgroup':Q1matgroup, 'Q2 Matgroup':Q2matgroup,'Q3 Matgroup':Q3matgroup,'Q4 Matgroup':Q4matgroup,'QU_Length_m':QU_Length, 'Box_type':Box_type, 'JS_C01_ID':inst_a, 'JS_C02_ID':inst_b}
        boxes_info.append(parameters_info)
    
    return boxes_info


def get_boxesfilteredbyPJ(ruta_archivo):
    
    allboxes_info=boxes_info_joint(ruta_archivo)

    parents_únicos=set()

    allboxes_info_filtered=[]

    for info in allboxes_info:
        parent=info['JS_ParentJointInstanceID']
        #print(parent)
        
        if parent not in parents_únicos:
            parents_únicos.add(parent)
            #print(parents_únicos)
            allboxes_info_filtered.append(info)
                
    return allboxes_info_filtered

def get_balconys(ruta):
    
    balcony=[]
    ifc=abrir_ifc(ruta)
    type='IfcWindow'
    filterbytype=ifc.by_type(type)
    
    for i in filterbytype:
        i_psets=ifcopenshell.util.element.get_psets(i)
        QU_Quantity = i_psets.get('QU_Quantity', {})
        EI_Elements_Identification = i_psets.get('EI_Elements Identification', {})
        EI_OpeningType = EI_Elements_Identification.get('EI_OpeningType', '')
        EI_HostComponentInstanceID=EI_Elements_Identification.get('EI_HostComponentInstanceID', '')
        QU_Height=QU_Quantity.get('QU_Height_m', '')
        if QU_Height>1.8:
            balcony.append({'EI_OpeningType':EI_OpeningType,'EI_HostComponentInstanceID':EI_HostComponentInstanceID,'QU_Height_m':QU_Height})
    return balcony

def get_huecosdepaso(ruta):
    
    doorspaso_info=[]
    ifc=abrir_ifc(ruta)
    type='IfcWall'
    filterbytype=ifc.by_type(type)
    
    for i in filterbytype:
        i_psets=ifcopenshell.util.element.get_psets(i)
        Pset_ProductRequirements=i_psets.get('Pset_ProductRequirements',{}) 
        EI_Elements_Identification=i_psets.get('EI_Elements Identification',{})   
        EI_HostComponentInstanceID=EI_Elements_Identification.get('EI_HostComponentInstanceID','')
        EI_OpeningType=EI_Elements_Identification.get('EI_OpeningType', '') 
        Category=Pset_ProductRequirements.get('Category', '')
        Params_info={'EI_HostComponentInstanceID':EI_HostComponentInstanceID,'EI_OpeningType':EI_OpeningType, 'Category':Category}
        if Category=='Doors':
                doorspaso_info.append(Params_info)
    return doorspaso_info

def get_doors(ruta):
    doors_info=[]
    ifc=abrir_ifc(ruta)
    type='IfcDoor'
    filterbytype=ifc.by_type(type)
    
    for i in filterbytype:
        i_psets=ifcopenshell.util.element.get_psets(i)
        Pset_ProductRequirements=i_psets.get('Pset_ProductRequirements',{}) 
        EI_Elements_Identification=i_psets.get('EI_Elements Identification',{})
        EI_HostComponentInstanceID=EI_Elements_Identification.get('EI_HostComponentInstanceID','')
        EI_OpeningType=EI_Elements_Identification.get('EI_OpeningType', '') 
        Category=Pset_ProductRequirements.get('Category', '')
        Params_info={'EI_HostComponentInstanceID':EI_HostComponentInstanceID,'EI_OpeningType':EI_OpeningType, 'Category':Category}
        doors_info.append(Params_info)
    return doors_info

def get_alldoors(ruta):
    alldoors_info=get_doors(ruta)+get_huecosdepaso(ruta)
    return alldoors_info


def nrtallopenings_byinstance(ruta):
    info_ifcwindow=get_balconys(ruta)
    alldoors_info=get_alldoors(ruta)
    recuento_balconerasporinstancia={}
    
    for balcony in info_ifcwindow:
        instancia=balcony['EI_HostComponentInstanceID']
        if instancia in recuento_balconerasporinstancia:
            recuento_balconerasporinstancia[instancia] += 1
        else:
            recuento_balconerasporinstancia[instancia] = 1
            
    for door in alldoors_info:
        instancia=door['EI_HostComponentInstanceID']
        if instancia in recuento_balconerasporinstancia:
            recuento_balconerasporinstancia[instancia] += 1
        else:
            recuento_balconerasporinstancia[instancia] = 1
    
    return recuento_balconerasporinstancia

def getboxesfilteredwithbalconies(ruta):
    herrajesmodelados=getmodeledconnections(ruta)
    balcony_instances_input=nrtallopenings_byinstance(ruta)
    allboxes_info_filtered_input=get_boxesfilteredbyPJ(ruta)
    boxesandbalconys=[]
    #PJ='C_FAC-0022_7203-00.0474'
    #print(balcony_instances_input[PJ])
    for i in allboxes_info_filtered_input:
        componente=i['JS_C01_ID']
        if i['Box_type']=="H.ST_Bottom":
            if componente!='' and componente in balcony_instances_input.keys():
                balconerasdelainstancia=balcony_instances_input[componente]
                i['nrbalconies']=balconerasdelainstancia
                boxesandbalconys.append(i)
                #print(i)
            else:
                i['nrbalconies']=0
                boxesandbalconys.append(i)
        else:
            i['nrbalconies']=0
            boxesandbalconys.append(i)
    return boxesandbalconys,herrajesmodelados

def getmodeledconnections(ruta):
    connections_info=[]    
    ifc=abrir_ifc(ruta)
    type='IfcBuildingElementProxy'
    pset='EI_Elements Identification'
    parameter='EI_Type'
    value='Connection'
    connections=list_elements(ifc, type, pset, parameter, value)
    for i in connections:
        i_psets = ifcopenshell.util.element.get_psets(i)        
        JS_Joint_Specification = i_psets.get('JS_Joint Specification', {})    
        EI_Interoperability=i_psets.get('EI_Interoperability', {})
        connectiontype_id = JS_Joint_Specification.get('JS_ConnectionTypeID', '')        
        parent_joint_id = JS_Joint_Specification.get('JS_ParentJointInstanceID', '')
        r_guid = EI_Interoperability.get('RevitGUID', '')      
        parameters_info={'RevitGUID':r_guid,'JS_ParentJointInstanceID': parent_joint_id, 'JS_ConnectionTypeID':connectiontype_id}
        connections_info.append(parameters_info)        
    
    recuento = {}
    
    for item in connections_info:
        parent_joint_id = item['JS_ParentJointInstanceID']
        connection_type_id = item['JS_ConnectionTypeID']
        
        if parent_joint_id in recuento:
            if connection_type_id in recuento[parent_joint_id]:
                recuento[parent_joint_id][connection_type_id] += 1
            else:
                recuento[parent_joint_id][connection_type_id] = 1
        else:
            recuento[parent_joint_id] = {connection_type_id: 1}
    
    resultados = []
    
    for parent_joint_id, connection_types in recuento.items():
        for connection_type_id, count in connection_types.items():
            resultados.append({
                'ParentJoint_id': parent_joint_id,
                'Connectiontype_id': connection_type_id,
                'nr_units': count
            })
    
    listconnectiontypesmodelados=[]
    for herraje in resultados:
        new_herraje={'Calculation Formula':'Fix value',
                     'Performance':herraje['nr_units'],
                     'connection_type':herraje['Connectiontype_id'],
                     'connectiongroup_type_id':'',
                     'is_modeled':'modeled',
                     'parentjoint_id':herraje['ParentJoint_id']}
        listconnectiontypesmodelados.append(new_herraje)
    
    return listconnectiontypesmodelados


#-------------------------------------------------------------------------- Funciones de exportación de las cajas a excel

def export_excel_infoallboxes(rutaifc):
    
    datos_export=boxes_info_joint(rutaifc)
            
    # Obtener el nombre del archivo IFC sin la extensión
    nombre_archivo_ifc = os.path.splitext(os.path.basename(rutaifc))[0]
    
    # Crear un DataFrame de pandas con los resultados
    df = pd.DataFrame(datos_export)
    
    # Obtener la ruta del directorio donde se encuentra el archivo IFC
    directorio_ifc = os.path.dirname(rutaifc)
    
    # Generar el nombre del archivo Excel con el sufijo "allboxes"
    excel_file_name = f"{nombre_archivo_ifc}_allboxes.xlsx"
    
    # Combinar la ruta del directorio del archivo IFC con el nombre del archivo Excel
    excel_file_path = os.path.join(directorio_ifc, excel_file_name)
    
    # Exportar el DataFrame a un archivo de Excel en la misma ruta que el archivo IFC
    df.to_excel(excel_file_path, index=False)  
    
    print(f"Resultados exportados a {excel_file_name} en la ruta: {directorio_ifc}")
    
def export_excel_infofilteredboxes(rutaifc):
    
    datos_export=get_boxesfilteredbyPJ(rutaifc)
            
    
    # Obtener el nombre del archivo IFC sin la extensión
    nombre_archivo_ifc = os.path.splitext(os.path.basename(rutaifc))[0]
    
    # Crear un DataFrame de pandas con los resultados
    df = pd.DataFrame(datos_export)
    
    # Obtener la ruta del directorio donde se encuentra el archivo IFC
    directorio_ifc = os.path.dirname(rutaifc)
    
    # Generar el nombre del archivo Excel con el sufijo 'filteredboxes'
    excel_file_name = f"{nombre_archivo_ifc}_filteredboxes.xlsx"
    
    # Combinar la ruta del directorio del archivo IFC con el nombre del archivo Excel
    excel_file_path = os.path.join(directorio_ifc, excel_file_name)
    
    # Exportar el DataFrame a un archivo de Excel en la misma ruta que el archivo IFC
    df.to_excel(excel_file_path, index=False)  
    
    print(f"Resultados exportados a {excel_file_name} en la ruta: {directorio_ifc}")

def export2excel(ruta):
    boxesinfo, herrajesmodelados = getboxesfilteredwithbalconies(ruta)
    
    # Obtener el nombre del archivo IFC sin la extensión
    nombre_archivo = os.path.splitext(os.path.basename(ruta))[0]
    
    # Crear un DataFrame de pandas con los resultados
    df = pd.DataFrame(boxesinfo)
    
    # Obtener la ruta del directorio donde se encuentra el archivo IFC
    directorio_ifc = os.path.dirname(ruta)
    
    # Generar el nombre del archivo Excel con el sufijo 'filteredboxesandnrbalconies'
    excel_file_name = f"{nombre_archivo}_filteredboxesandnrbalconies.xlsx"
    
    # Combinar la ruta del directorio del archivo IFC con el nombre del archivo Excel
    ruta_excel = os.path.join(directorio_ifc, excel_file_name)
    
    # Exportar el DataFrame a un archivo de Excel en la misma ruta que el archivo IFC
    df.to_excel(ruta_excel, index=False)
    
    print(f'Exportación a {excel_file_name} completada en la ruta: {directorio_ifc}')

#--------------------------------------------------------------------------- Funciones búsqueda datos para generar BOQ Joints y Connections

import requests


class Airtable:
    def __init__(self, token, base_id):
        self.token = token
        self.base_id = base_id

    def list(self, table, max_records=None, view='API',fields=None,filter=None,sort=None):
        url = f'https://api.airtable.com/v0/{self.base_id}/{table}'
        params = {
            'maxRecords': max_records,
            'view': view
        }
        if filter is not None:
            params['filterByFormula'] = filter
        if fields is not None:
            params['fields[]'] = fields
        if sort is not None:
            params['sort[]'] = sort
        data = []

        while True:
            response = requests.get(url, params=params, headers={'Authorization': 'Bearer ' + self.token})
            #print (response.url)
            response_data = response.json()
            data.extend(response_data.get('records', []))
            offset = response_data.get('offset', None)
            #print (offset)
            if not offset:
                break
            params['offset'] = offset
            
        return data
        # example with custom filter: at.list('Contracts',filter="AND({Contract Type}='Obra' , {Project}='5')")

    def insert(self, table, data):
        url = f'https://api.airtable.com/v0/{self.base_id}/{table}'
        response = requests.post(url, json={'fields': data}, headers={'Authorization': 'Bearer ' + self.token})
        return response.json()
    
    
def airtable_conection(api_key,base_id):
    connection=Airtable(api_key,base_id)    
    return connection

jointsplayground_connect = airtable_conection(adri_jointsplayground_api_key,joints_base_id)
# materials_connect = airtable_conection(adri_materiales_api_key,materiales_base_id)
joints3playground_connect=airtable_conection(adri_joints3playground_api_key,materiallayers_base_id)

def jointsplayground_connectiongroup():    
    resultados_cgtype = jointsplayground_connect.list(cgtype_table,view='AM', fields=['cgtype_id','Description','api_ConnectionGroup_Class','RL_ConnectionGroupType_ConnectionType_api'])
    connectiongroup_types = []
    for i in resultados_cgtype:
        connectiongroup_types.append(i['fields'])
    return connectiongroup_types

def connectiontype_of_connectiongroup(connectiongroup_type):
    resultados_rl_cgtype_ctype = jointsplayground_connect.list(rl_cgtype_ctype_table, view='AMG_Export', fields=['connectiongroup_type_id','connection_type','Performance','Calculation Formula'],filter="SEARCH('"+connectiongroup_type+"', {connectiongroup_type_id})")
    connection_types = []    
    for i in resultados_rl_cgtype_ctype:
        connection_types.append(i['fields'])
    return connection_types
    #tipo de respuesta: [{'Calculation Formula': 'Fix value', 'Performance': 3, 'connection_type': 'H_T3-0003', 'connectiongroup_type_id': 'CG_0004'}, {'Calculation Formula': 'Length * performance', 'Performance': 5, 'connection_type': 'H_C1-0023', 'connectiongroup_type_id': 'CG_0004'}]

def connectionlayers_of_connectiontype(connection_type):
    resultados_connectionlayers=jointsplayground_connect.list(clayers_table,fields=['connection_type_code','material_id','Performance','Calculation Formula','Current_material_cost','Units'],filter="SEARCH('"+connection_type+"', {connection_type_code})")
    clayers=[]
    for i in resultados_connectionlayers:
        clayers.append(i['fields'])
    return(clayers)
    #tipo de respuesta: [{'Performance': 1, 'Calculation Formula': 'Fix value', 'connection_type_code': 'H_T3-0003', 'material_id': 'MBRA0214'}, {'Performance': 60, 'Calculation Formula': 'Fix value', 'connection_type_code': 'H_T3-0003', 'material_id': 'MFIX0578'}]

    
def jlayers(joint):
    resultados_jlayers=jointsplayground_connect.list(jointlayers_table,fields=['joint_type_code','material_id','Performance','Calculation Formula','Current_material_cost'],filter="SEARCH('"+joint+"',{joint_type_code})")
    joints=[]
    for i in resultados_jlayers:
        joints.append(i['fields'])
    return joints
    #tipo de respuesta: {'Performance': 1.05, 'Calculation Formula': 'Length * performance', 'joint_type_code': 'J_0351', 'material_id': 'MJNT4141', 'Current_material_cost': [5.81]}


def rl_cgtype_ctype():
    resultados_rl_cgtype_ctype = jointsplayground_connect.list(rl_cgtype_ctype_table, view='AMG_Export', fields=['connectiongroup_type_id','connection_type','Performance','Calculation Formula','is_modeled'])
    connection_types = []    
    for i in resultados_rl_cgtype_ctype:
        connection_types.append(i['fields'])
    return connection_types
    #tipo de respuesta: [{'Calculation Formula': 'Fix value', 'Performance': 3, 'connection_type': 'H_T3-0003', 'connectiongroup_type_id': 'CG_0004'}, {'Calculation Formula': 'Length * performance', 'Performance': 5, 'connection_type': 'H_C1-0023', 'connectiongroup_type_id': 'CG_0004'}]

def connectionlayers():
    resultados_connectionlayers=jointsplayground_connect.list(clayers_table,fields=['connection_type_code','material_id','Performance','Calculation Formula','Current_material_cost','Units','Description (from Material)','Fase', 'is_modeled'])
    clayers=[]
    for i in resultados_connectionlayers:
        clayers.append(i['fields'])
    return(clayers)
    #tipo de respuesta: [{'Performance': 1, 'Calculation Formula': 'Fix value', 'connection_type_code': 'H_T3-0003', 'material_id': 'MBRA0214'}, {'Performance': 60, 'Calculation Formula': 'Fix value', 'connection_type_code': 'H_T3-0003', 'material_id': 'MFIX0578'}]

def alljlayers():
    resultados_jlayers=jointsplayground_connect.list(jointlayers_table,fields=['api_id','material_id','Performance','Calculation Formula','Current_material_cost','Long_Name (from Material)','Fase','units'])
    joints=[]
    for i in resultados_jlayers:
        joints.append(i['fields'])
    return joints    
    # tipo de respuesta: {'Performance': 1.05, 'Calculation Formula': 'Length * performance', 'joint_type_code': 'J_0351', 'material_id': 'MJNT4141', 'Current_material_cost': [5.81]}
    
def allmatlayers():
    resultados_matlayers=joints3playground_connect.list(materiallayers_table,view='API',fields=['api_id','material_id','Performance','Calculation Formula','Current_material_cost','Long_Name (from Material)','Fase','units'])
    matlayers=[]
    for i in resultados_matlayers:
        matlayers.append(i['fields'])
    return matlayers
    #tipo de respuesta: {'Performance': 1.05, 'calculation_formula': 'Length * performance', 'Fase': 'Onsite for Assembly', 'material_description': ['MJNT0847-Cinta monoadhesiva universal ROTHOBLAAS FLEXI60 60x25000mm -Poliacrilato -60   -ml'], 'api_material_group': 'CORE_17', 'api_Material': 'MJNT0847', 'current_material_cost': [0.54116]}



def connectiongroup_costcalculator(parentid,connectiongroup_type, long, openings, airtable_rlcgctype_data, airtable_clayers):
    finalclayers=[]    
    filteredconnectiontypes=[]
    for i in airtable_rlcgctype_data:
        if connectiongroup_type in i['connectiongroup_type_id'] and connectiongroup_type!='':
            i['parentjoint_id']=parentid
            filteredconnectiontypes.append(i)
        
    connectiongroup_cost = 0
    for connectiontype in filteredconnectiontypes:
        connectiontype_id=connectiontype['connection_type']
        connectiontype_calcform=connectiontype['Calculation Formula']
        connectiontype_performance=connectiontype['Performance']          
        allclayers=airtable_clayers
        filteredclayers=[]
        
        for i in allclayers:
            if connectiontype_id in i['connection_type_code']:
                i['parentjoint_id']=parentid
                filteredclayers.append(i)
                
        for clayer in filteredclayers:
            material_performance=clayer['Performance']
            material_formula=clayer['Calculation Formula']
            material_sku=clayer['material_id']
            material_cost=clayer.get('Current_material_cost',[0])[0]
            material_unit=clayer['Units'][0]                     
            
            if material_formula=='Fix value':
                if material_unit=='U':
                    if connectiontype_calcform=='Length * performance':
                        quantity=math.ceil(material_performance*connectiontype_performance*long)
                        connectiongroup_cost=connectiongroup_cost+material_cost*quantity
                        clayer['quantity']=quantity
                        clayer['layer_cost']=material_cost*quantity
                    elif connectiontype_calcform=='Opening * performance':
                        quantity=material_performance*connectiontype_performance*openings
                        connectiongroup_cost=connectiongroup_cost+material_cost*quantity                        
                        clayer['quantity']=quantity  
                        clayer['layer_cost']=material_cost*quantity                      
                    else:
                        quantity=material_performance*connectiontype_performance
                        connectiongroup_cost=connectiongroup_cost+material_cost*quantity                        
                        clayer['quantity']=quantity
                        clayer['layer_cost']=material_cost*quantity
                else:
                    if connectiontype_calcform=='Length * performance':
                        quantity=material_performance*connectiontype_performance*long
                        connectiongroup_cost=connectiongroup_cost+material_cost*quantity                        
                        clayer['quantity']=quantity
                        clayer['layer_cost']=material_cost*quantity
                    elif connectiontype_calcform=='Opening * performance':
                        quantity=material_performance*connectiontype_performance*openings
                        connectiongroup_cost=connectiongroup_cost+material_cost*quantity                        
                        clayer['quantity']=quantity
                        clayer['layer_cost']=material_cost*quantity
                    else:
                        quantity=material_performance*connectiontype_performance
                        connectiongroup_cost=connectiongroup_cost+material_cost*quantity                        
                        clayer['quantity']=quantity
                        clayer['layer_cost']=material_cost*quantity
            finalclayers.append(clayer)
    return connectiongroup_cost,finalclayers    

def inferredandmodeled_advanced_connectiongroup_costcalculator(parentid,connectiongroup_type, long, openings, airtable_rlcgctype_data, airtable_clayers,herrajesmodelados): #calcula el coste y la lista teniendo en cuenta para cada parent joint si tiene elementos modelados o no
    inferredconnectiontypes=[]
    modeledconnectiontypes=herrajesmodelados
    finalconnectiontypes=[]
    
    inferredandmodeledclayers=[]

    listadeparentsconalgunherrajemodelado=[]
    
    for i in modeledconnectiontypes:
        if i['parentjoint_id'] not in listadeparentsconalgunherrajemodelado:
            listadeparentsconalgunherrajemodelado.append(i['parentjoint_id'])
        
    
    for line in airtable_rlcgctype_data:
        
        if connectiongroup_type in line['connectiongroup_type_id'] and connectiongroup_type!='':
            line['parentjoint_id']=parentid                  
            inferredconnectiontypes.append(line)
            
    if inferredconnectiontypes==[]:
        for herraje in modeledconnectiontypes:
            if parentid==herraje['parentjoint_id']:
                finalconnectiontypes.append(herraje)
        
    
    recuentodeconnectiontypes=[]
            
    for connectiontype in inferredconnectiontypes:
        if connectiontype['is_modeled'][0]=='Yes':
            if connectiontype['parentjoint_id'] not in listadeparentsconalgunherrajemodelado:
                finalconnectiontypes.append(connectiontype)
            for herraje in modeledconnectiontypes:
                if herraje['parentjoint_id']==connectiontype['parentjoint_id'] and herraje['connection_type']==connectiontype['connection_type']:
                #si coinciden el parent y el connection type
                    if herraje['connection_type'] not in recuentodeconnectiontypes:
                        newconnectiontype=connectiontype.copy()
                        newconnectiontype['Calculation Formula']='Fix value' 
                        newconnectiontype['Performance']=herraje['Performance']  
                        newconnectiontype['is_modeled']='modeled' 
                        recuentodeconnectiontypes.append(herraje['connection_type'])
                        finalconnectiontypes.append(newconnectiontype)
                if herraje['parentjoint_id']==connectiontype['parentjoint_id'] and herraje['connection_type']!=connectiontype['connection_type']:
                #si coinciden el parent y pero el connectiontype es diferente    
                    if herraje['connection_type'] not in recuentodeconnectiontypes:
                        newconnectiontype={}
                        newconnectiontype['Calculation Formula']='Fix value' 
                        newconnectiontype['Performance']=herraje['Performance'] 
                        newconnectiontype['connection_type']=herraje['connection_type']
                        newconnectiontype['connectiongroup_type_id']='' 
                        newconnectiontype['is_modeled']='modeled' 
                        newconnectiontype['parentjoint_id']=herraje['parentjoint_id']
                        recuentodeconnectiontypes.append(herraje['connection_type'])
                        finalconnectiontypes.append(newconnectiontype)
                
                    
        if connectiontype['is_modeled'][0]=='No':
            finalconnectiontypes.append(connectiontype)
                        
    connectiongroup_cost = 0
    for connectiontype in finalconnectiontypes:
        connectiontype_id=connectiontype['connection_type']
        connectiontype_calcform=connectiontype['Calculation Formula']
        connectiontype_performance=connectiontype['Performance']          
        allclayers=airtable_clayers
        filteredclayers=[]
        
        for i in allclayers:
            if connectiontype_id in i['connection_type_code']:
                i['parentjoint_id']=parentid
                filteredclayers.append(i)
                
        for clayer in filteredclayers:
            material_performance=clayer['Performance']
            material_formula=clayer['Calculation Formula']
            material_sku=clayer['material_id']
            material_cost=clayer.get('Current_material_cost',[0])[0]
            material_unit=clayer['Units'][0]                     
            
            if material_formula=='Fix value':
                if material_unit=='U':
                    if connectiontype_calcform=='Length * performance':
                        quantity=math.ceil(material_performance*connectiontype_performance*long)
                        connectiongroup_cost=connectiongroup_cost+material_cost*quantity
                        clayer['quantity']=quantity
                        clayer['layer_cost']=material_cost*quantity
                    elif connectiontype_calcform=='Opening * performance':
                        quantity=material_performance*connectiontype_performance*openings
                        connectiongroup_cost=connectiongroup_cost+material_cost*quantity                        
                        clayer['quantity']=quantity
                        clayer['layer_cost']=material_cost*quantity                        
                    else:
                        quantity=material_performance*connectiontype_performance
                        connectiongroup_cost=connectiongroup_cost+material_cost*quantity                        
                        clayer['quantity']=quantity
                        clayer['layer_cost']=material_cost*quantity
                else:
                    if connectiontype_calcform=='Length * performance':
                        quantity=material_performance*connectiontype_performance*long
                        connectiongroup_cost=connectiongroup_cost+material_cost*quantity                        
                        clayer['quantity']=quantity
                        clayer['layer_cost']=material_cost*quantity
                    elif connectiontype_calcform=='Opening * performance':
                        quantity=material_performance*connectiontype_performance*openings
                        connectiongroup_cost=connectiongroup_cost+material_cost*quantity                        
                        clayer['quantity']=quantity
                        clayer['layer_cost']=material_cost*quantity
                    else:
                        quantity=material_performance*connectiontype_performance
                        connectiongroup_cost=connectiongroup_cost+material_cost*quantity                        
                        clayer['quantity']=quantity
                        clayer['layer_cost']=material_cost*quantity
            inferredandmodeledclayers.append(clayer)        
            
    return connectiongroup_cost, inferredandmodeledclayers 

def inferrednotmodeled_connectiongroup_costcalculator(parentid,connectiongroup_type, long, openings, airtable_rlcgctype_data, airtable_clayers,herrajesmodelados): #calcula el coste y la lista teniendo en cuenta para cada parent joint si tiene elementos modelados o no
    inferrednotmodeledclayers=[]    
    inferredconnectiontypes=[]
    inferrednotmodeledconnectiontypes=[]
        
    parent_connectiontype={}
    
    recuento={}
    
    for herraje in herrajesmodelados:
        ParentJoint_id=herraje['parentjoint_id']
        Connectiontype_id=herraje['connection_type']
        if ParentJoint_id in parent_connectiontype:
            parent_connectiontype[ParentJoint_id].append(Connectiontype_id)
        else:
            parent_connectiontype[ParentJoint_id]=[Connectiontype_id]
    
    # parent_connectiontype_formateado=[{'ParentJoint_id':key,'Connectiontype_id':value} for key,value in parent_connectiontype.items()]
                    
    for line in airtable_rlcgctype_data:
        
        if connectiongroup_type in line['connectiongroup_type_id'] and connectiongroup_type!='':
            line['parentjoint_id']=parentid                  
            inferredconnectiontypes.append(line)
    
    for line in inferredconnectiontypes:        
        if line['is_modeled'][0]=='No':
            inferrednotmodeledconnectiontypes.append(line)
    
    connectiongroup_cost = 0
    for connectiontype in inferrednotmodeledconnectiontypes:
        connectiontype_id=connectiontype['connection_type']
        connectiontype_calcform=connectiontype['Calculation Formula']
        connectiontype_performance=connectiontype['Performance']          
        allclayers=airtable_clayers
        filteredclayers=[]
        
        for i in allclayers:
            if connectiontype_id in i['connection_type_code']:
                i['parentjoint_id']=parentid
                filteredclayers.append(i)
                
        for clayer in filteredclayers:
            material_performance=clayer['Performance']
            material_formula=clayer['Calculation Formula']
            material_sku=clayer['material_id']
            material_cost=clayer.get('Current_material_cost',[0])[0]
            material_unit=clayer['Units'][0]                     
            
            if material_formula=='Fix value':
                if material_unit=='U':
                    if connectiontype_calcform=='Length * performance':
                        quantity=math.ceil(material_performance*connectiontype_performance*long)
                        connectiongroup_cost=connectiongroup_cost+material_cost*quantity
                        clayer['quantity']=quantity
                        clayer['layer_cost']=material_cost*quantity
                    elif connectiontype_calcform=='Opening * performance':
                        quantity=material_performance*connectiontype_performance*openings
                        connectiongroup_cost=connectiongroup_cost+material_cost*quantity                        
                        clayer['quantity']=quantity
                        clayer['layer_cost']=material_cost*quantity                        
                    else:
                        quantity=material_performance*connectiontype_performance
                        connectiongroup_cost=connectiongroup_cost+material_cost*quantity                        
                        clayer['quantity']=quantity
                        clayer['layer_cost']=material_cost*quantity
                else:
                    if connectiontype_calcform=='Length * performance':
                        quantity=material_performance*connectiontype_performance*long
                        connectiongroup_cost=connectiongroup_cost+material_cost*quantity                        
                        clayer['quantity']=quantity
                        clayer['layer_cost']=material_cost*quantity
                    elif connectiontype_calcform=='Opening * performance':
                        quantity=material_performance*connectiontype_performance*openings
                        connectiongroup_cost=connectiongroup_cost+material_cost*quantity                        
                        clayer['quantity']=quantity
                        clayer['layer_cost']=material_cost*quantity
                    else:
                        quantity=material_performance*connectiontype_performance
                        connectiongroup_cost=connectiongroup_cost+material_cost*quantity                        
                        clayer['quantity']=quantity
                        clayer['layer_cost']=material_cost*quantity
            inferrednotmodeledclayers.append(clayer)
    return inferrednotmodeledclayers

def realmodeledconnections_costcalculator(herrajesmodelados, airtable_clayers):
    realmodeled_clayers=[]
    connectiontypes=herrajesmodelados
    # for herraje in herrajesmodelados:
    #     connectiontype={}
    #     connectiontype['Calculation Formula']='Fix value'
    #     connectiontype['Performance']=herraje['Performance']
    #     connectiontype['connection_type']=herraje['connection_type']
    #     connectiontype['connectiongroup_type_id']=''
    #     connectiontype['is_modeled']=''
    #     connectiontype['parentjoint_id']='NoParentAssociated'
    #     if connectiontype['connection_type']!='':
    #         connectiontypes.append(connectiontype)
    
    for connectiontype in connectiontypes:
        connectiontype_id=connectiontype['connection_type']
        connectiontype_calcform=connectiontype['Calculation Formula']
        connectiontype_performance=connectiontype['Performance']          
        allclayers=airtable_clayers
        filteredclayers=[]
        
        for i in allclayers:
            if connectiontype_id == i['connection_type_code'] and connectiontype_id!='':
                new_layer=i.copy()
                new_layer['parentjoint_id']=connectiontype.get('parentjoint_id','NoParentAssociated')
                filteredclayers.append(new_layer)
                
        for clayer in filteredclayers:
            material_performance=clayer['Performance']
            material_formula=clayer['Calculation Formula']            
            material_cost=clayer.get('Current_material_cost',[0])[0]
            material_unit=clayer['Units'][0]                     
            
            if material_formula=='Fix value':
                if material_unit=='U':                                                                          
                    quantity=material_performance*connectiontype_performance                                            
                    clayer['quantity']=quantity
                    clayer['layer_cost']=material_cost*quantity
                else:                    
                    quantity=material_performance*connectiontype_performance                                            
                    clayer['quantity']=quantity
                    clayer['layer_cost']=material_cost*quantity
            realmodeled_clayers.append(clayer)
    
    return realmodeled_clayers 
 
def jointtype_costcalculator (parentid,joint, long, airtable_jlayers):
    alljoint_layers=airtable_jlayers
    filteredjlayers=[]
    for i in alljoint_layers:
        if joint in i['api_id'] and joint!='':
            i['parentjoint_id']=parentid
            filteredjlayers.append(i)        
    joint_cost=0
    for layer in filteredjlayers:
        material_sku=layer['material_id']
        layer_performance=layer['Performance']
        layer_formula=layer['Calculation Formula']
        layer_currentmaterialcost=layer.get('Current_material_cost',[0])[0]
        if layer_formula=='Length * performance':
            quantity= layer_performance*long           
            layer_cost=layer_currentmaterialcost*quantity 
        elif layer_formula=='Fix value':
            quantity=layer_performance
            layer_cost=layer_currentmaterialcost*quantity  
        layer['quantity']=quantity 
        layer['layer_cost']=layer_cost       
        joint_cost=joint_cost+layer_cost
    # print(joint_cost)
    return joint_cost,filteredjlayers
          
# def costeunion (parentid,joint, connectiongroup_type, long, openings,airtable_rlcgctype_data, airtable_clayers, airtable_jlayers):
    costejoint, joint_materials=jointtype_costcalculator(parentid, joint, long, airtable_jlayers)
    costeconnection, connection_materials=connectiongroup_costcalculator(parentid, connectiongroup_type,long,openings,airtable_rlcgctype_data, airtable_clayers)
    coste=costejoint+costeconnection
    return coste, joint_materials, connection_materials

def guardarxlsfromIFC(listaconcoste, listamateriales, listaherrajes, rutaifc):
    nombrearchivo=os.path.splitext(os.path.basename(rutaifc))[0]
    df = pd.DataFrame(listaconcoste)
    ruta_excel = os.path.join(os.path.dirname(rutaifc), f'{nombrearchivo}.xlsx')
    df.to_excel(ruta_excel, index=False)    
    
    excel_file = openpyxl.load_workbook(ruta_excel)  
    
    excel_file.active.title="Parentwithcost"  
   
    nueva_hoja_materiales = excel_file.create_sheet(title="Lista mat joints")
    if listamateriales:
        keys = list(listamateriales[0].keys())
        nueva_hoja_materiales.append(keys)
        for material in listamateriales:
            row_data = [material[key][0] if isinstance(material.get(key, ''), list) else material.get(key, '') for key in keys]
            nueva_hoja_materiales.append(row_data)
    
    # Crea una hoja "Lista herrajes" y guarda los datos de listaherrajes
    nueva_hoja_herrajes = excel_file.create_sheet(title="Lista herrajes")
    if listaherrajes:
        keys = list(listaherrajes[0].keys())
        nueva_hoja_herrajes.append(keys)
        for herraje in listaherrajes:
            row_data = [herraje[key][0] if isinstance(herraje.get(key, ''), list) else herraje.get(key, '') for key in keys]
            nueva_hoja_herrajes.append(row_data)
    
    # Guarda el archivo Excel actualizado
    excel_file.save(ruta_excel)
    print('Archivo guardado con éxito')
    

def leer_archivo_xlsx(ruta):
    # Leer el archivo xlsx
    df = pd.read_excel(ruta, header=None)  # No asumir que la primera fila es un encabezado

    # Verificar si hay datos en la primera columna
    primera_columna = df.iloc[:, 0]
    ultima_fila = primera_columna.last_valid_index()

    if ultima_fila is None:
        raise ValueError("No se encontraron datos en la primera columna.")

    # Filtrar el DataFrame hasta la última fila con algún valor en la primera columna
    df = df.loc[:ultima_fila]

    # Crear una lista de diccionarios con los datos de cada fila
    lista_resultado = []
    for index, fila in df.iterrows():
        diccionario_fila = {
            'JS_ParentJointInstanceID': fila[0] if pd.notna(fila[0]) else '',
            'JS_JointTypeID': fila[1] if pd.notna(fila[1]) else '',
            'JS_ConnectionGroupTypeID': fila[2] if pd.notna(fila[2]) else '',
            'Core Matgroup': fila[3] if pd.notna(fila[3]) else '',
            'Q1 Matgroup': fila[4] if pd.notna(fila[4]) else '',
            'Q2 Matgroup': fila[5] if pd.notna(fila[5]) else '',
            'Q3 Matgroup': fila[6] if pd.notna(fila[6]) else '',
            'Q4 Matgroup': fila[7] if pd.notna(fila[7]) else '',
            'QU_Length_m': fila[8] if pd.notna(fila[8]) else '',
            'nrbalconies': fila[9] if pd.notna(fila[9]) else ''
        }
        lista_resultado.append(diccionario_fila)

    return lista_resultado

def matgroup_costcalculator (parentid,matgroup, long, airtable_matlayers):
    matgroupelements=matgroup.split(',')        
    filteredmatlayers=[]
    for i in airtable_matlayers:
        for matgroup in matgroupelements:
            if matgroup in i['api_id'] and matgroup!='':
                i['parentjoint_id']=parentid
                filteredmatlayers.append(i)        
    matgroup_cost=0
    for layer in filteredmatlayers:
        material_sku=layer['material_id']
        layer_performance=layer['Performance']
        layer_formula=layer['Calculation Formula']
        layer_currentmaterialcost=layer.get('Current_material_cost',[0])[0]
        if layer_formula=='Length * performance':
            quantity= layer_performance*long           
            layer_cost=layer_currentmaterialcost*quantity 
        elif layer_formula=='Fix value':
            quantity=layer_performance
            layer_cost=layer_currentmaterialcost*quantity  
        layer['quantity']=quantity
        layer['layer_cost']=layer_cost         
        matgroup_cost=matgroup_cost+layer_cost
    # print(joint_cost)
    return matgroup_cost,filteredmatlayers

def costeunionJ3 (parentid,joint, connectiongroup_type, corematgroup, Q1magtroup, Q2matgroup, Q3matgroup, Q4matgroup, long, openings,airtable_rlcgctype_data, airtable_clayers, airtable_jlayers, airtable_matlayers,herrajesmodelados,inputcalculoconexion):
    
       
    costejoint, joint_materials=jointtype_costcalculator(parentid,joint, long, airtable_jlayers)
    costecorematgroup, corematgrouplayers=matgroup_costcalculator(parentid,corematgroup, long, airtable_matlayers)
    costeQ1matgroup, Q1matgrouplayers=matgroup_costcalculator(parentid,Q1magtroup, long, airtable_matlayers)
    costeQ2matgroup, Q2matgrouplayers=matgroup_costcalculator(parentid,Q2matgroup, long, airtable_matlayers)
    costeQ3matgroup, Q3matgrouplayers=matgroup_costcalculator(parentid,Q3matgroup, long, airtable_matlayers)
    costeQ4matgroup, Q4matgrouplayers=matgroup_costcalculator(parentid,Q4matgroup, long, airtable_matlayers)
    if inputcalculoconexion == 'a':
        # print('Se reportarán los datos inferidos en las cajas')
        costeconnection, connection_materials=connectiongroup_costcalculator(parentid,connectiongroup_type,long,openings,airtable_rlcgctype_data, airtable_clayers)
        coste=costejoint+costeconnection+costecorematgroup+costeQ1matgroup+costeQ2matgroup+costeQ3matgroup+costeQ4matgroup
    elif inputcalculoconexion=='b':
        # print('Se comprobará para cada parent si tiene herrajes modelados. Si hay algo modelado será reportado junto con lo inferido que tenga tipificado "is_modeled=No"')
        costeconnection, connection_materials=inferredandmodeled_advanced_connectiongroup_costcalculator(parentid,connectiongroup_type,long,openings,airtable_rlcgctype_data, airtable_clayers,herrajesmodelados)
        coste=costejoint+costeconnection+costecorematgroup+costeQ1matgroup+costeQ2matgroup+costeQ3matgroup+costeQ4matgroup
    elif inputcalculoconexion=='c':
        # print("""De las cajas únicamente se reportará todo lo que esté marcado como "is_modeled=No".
        #       La pestaña "Parentwithcost" no mostrará el coste correcto. Revisar coste en Listamateriales y Listaherrajes""")
        connection_inferred_materials=inferrednotmodeled_connectiongroup_costcalculator(parentid,connectiongroup_type,long,openings,airtable_rlcgctype_data, airtable_clayers,herrajesmodelados)
        # connection_modeled_materials=realmodeledconnections_costcalculator(herrajesmodelados,airtable_clayers)
        coste=''
        connection_materials=connection_inferred_materials
    else:
        pass
        
    # matgroup_materials=corematgrouplayers+Q1matgrouplayers+Q2matgrouplayers+Q3matgrouplayers+Q4matgrouplayers
    return coste, joint_materials, connection_materials, corematgrouplayers,Q1matgrouplayers,Q2matgrouplayers,Q3matgrouplayers,Q4matgrouplayers

def boqfromlistofparentsJ3(parents,herrajesmodelados,inputcalculoconexion):
    airtable_rlcgctype_data=rl_cgtype_ctype()
    airtable_clayers=connectionlayers()
    airtable_jlayers=alljlayers()
    airtable_matgrouplayers=allmatlayers()
    listaconcoste=[]
    listamateriales=[]    
    listadeherrajes=[]
    
    if inputcalculoconexion=='c':
        connection_modeled_materials=realmodeledconnections_costcalculator(herrajesmodelados,airtable_clayers)
        for i in connection_modeled_materials:
            i['cgtype_id']=''
            listadeherrajes.append(i)
        
    
    for parent in parents:
        coste_parent, materiales, herrajes, corematerials,Q1materials,Q2materials,Q3materials,Q4materials=costeunionJ3(parent['JS_ParentJointInstanceID'],parent['JS_JointTypeID'],parent['JS_ConnectionGroupTypeID'],parent['Core Matgroup'],parent['Q1 Matgroup'],parent['Q2 Matgroup'],parent['Q3 Matgroup'],parent['Q4 Matgroup'],parent['QU_Length_m'],parent['nrbalconies'],airtable_rlcgctype_data, airtable_clayers, airtable_jlayers,airtable_matgrouplayers,herrajesmodelados,inputcalculoconexion)
        parent['coste_parent']=coste_parent
        listaconcoste.append(parent)
        for material in materiales:
            if material['api_id']!='':
                new_material=material.copy()
                # material['parentjoint_id']=parent['JS_ParentJointInstanceID']
                listamateriales.append(new_material)
        for herraje in herrajes:
            # Crea una copia independiente de 'herraje'
            nuevo_herraje = herraje.copy()
            
            # Agrega los parámetros adicionales
            # nuevo_herraje['parentjoint_id'] = parent['JS_ParentJointInstanceID']
            nuevo_herraje['cgtype_id'] = parent['JS_ConnectionGroupTypeID']
            
            # Agrega la copia a la lista 'listadeherrajes'
            listadeherrajes.append(nuevo_herraje)
        for corematerial in corematerials:            
            nuevo_corematerial = corematerial.copy()
                        
            # nuevo_corematerial['parentjoint_id'] = parent['JS_ParentJointInstanceID']
                                   
            listamateriales.append(nuevo_corematerial)
            
        for Q1material in Q1materials:            
            nuevo_Q1material = Q1material.copy()
                        
            # nuevo_Q1material['parentjoint_id'] = parent['JS_ParentJointInstanceID']
                      
            listamateriales.append(nuevo_Q1material)
            
            
        for Q2material in Q2materials:            
            nuevo_Q2material = Q2material.copy()
                        
            # nuevo_Q2material['parentjoint_id'] = parent['JS_ParentJointInstanceID']
                      
            listamateriales.append(nuevo_Q2material)
            
            
        for Q3material in Q3materials:            
            nuevo_Q3material = Q3material.copy()
                        
            # nuevo_Q3material['parentjoint_id'] = parent['JS_ParentJointInstanceID']
                      
            listamateriales.append(nuevo_Q3material)
            
        for Q4material in Q4materials:            
            nuevo_Q4material = Q4material.copy()
                        
            # nuevo_Q4material['parentjoint_id'] = parent['JS_ParentJointInstanceID']
                      
            listamateriales.append(nuevo_Q4material)
                        
            
    return listaconcoste,listamateriales,listadeherrajes

def guardarxlsfromxls(listaconcoste, listamateriales, listaherrajes, ruta_excel):       
    # Abre el archivo Excel
    excel_file = openpyxl.load_workbook(ruta_excel)
    
    # Crea una nueva hoja "Parent with cost" y guarda los datos de listaconcoste
    nueva_hoja = excel_file.create_sheet(title="Parent with cost")
    if listaconcoste:
        keys = list(listaconcoste[0].keys())
        nueva_hoja.append(keys)
        for elemento in listaconcoste:
            row_data = [elemento[key] for key in keys]
            nueva_hoja.append(row_data)
        
    # Crea una hoja "Lista mat joints" y guarda los datos de listamateriales
    nueva_hoja_materiales = excel_file.create_sheet(title="Lista mat joints")
    if listamateriales:
        keys = list(listamateriales[0].keys())
        nueva_hoja_materiales.append(keys)
        for material in listamateriales:
            row_data = [material[key][0] if isinstance(material.get(key, ''), list) else material.get(key, '') for key in keys]
            nueva_hoja_materiales.append(row_data)
    
    # Crea una hoja "Lista herrajes" y guarda los datos de listaherrajes
    nueva_hoja_herrajes = excel_file.create_sheet(title="Lista herrajes")
    if listaherrajes:
        keys = list(listaherrajes[0].keys())
        nueva_hoja_herrajes.append(keys)
        for herraje in listaherrajes:
            row_data = [herraje[key][0] if isinstance(herraje.get(key, ''), list) else herraje.get(key, '') for key in keys]
            nueva_hoja_herrajes.append(row_data)
    
    # Guarda el archivo Excel actualizado
    excel_file.save(ruta_excel)
    print('Archivo guardado con éxito')


#-----------------------------------------------------PRUEBAS-------------------------------------------------
#-------------------------------------------------------------------------------------------------------------
#-------------------------------------------------------------------------------------------------------------

# def WIP_inferredandmodeled_advanced_connectiongroup_costcalculator(parentid,connectiongroup_type, long, openings, airtable_rlcgctype_data, airtable_clayers,herrajesmodelados): 
    # inferredconnectiontypes=[]
    # modeledconnectiontypes=herrajesmodelados
    # finalconnectiontypes=[]
    
    # inferredandmodeledclayers=[]

    # listadeparentsconalgunherrajemodelado=[]
    
    # for i in modeledconnectiontypes:
    #     if i['parentjoint_id'] not in listadeparentsconalgunherrajemodelado:
    #         listadeparentsconalgunherrajemodelado.append(i['parentjoint_id'])
        
    
    # for line in airtable_rlcgctype_data:
        
    #     if connectiongroup_type in line['connectiongroup_type_id'] and connectiongroup_type!='':
    #         line['parentjoint_id']=parentid                  
    #         inferredconnectiontypes.append(line)
            
    # if inferredconnectiontypes==[]:
    #     for herraje in modeledconnectiontypes:
    #         if parentid==herraje['parentjoint_id']:
    #             finalconnectiontypes.append(herraje)
        
    
    # recuentodeconnectiontypes=[]
            
    # for connectiontype in inferredconnectiontypes:
    #     if connectiontype['is_modeled'][0]=='Yes':
    #         if connectiontype['parentjoint_id'] not in listadeparentsconalgunherrajemodelado:
    #             finalconnectiontypes.append(connectiontype)
    #         for herraje in modeledconnectiontypes:
    #             if herraje['parentjoint_id']==connectiontype['parentjoint_id'] and herraje['connection_type']==connectiontype['connection_type']:
    #             #si coinciden el parent y el connection type
    #                 if herraje['connection_type'] not in recuentodeconnectiontypes:
    #                     newconnectiontype=connectiontype.copy()
    #                     newconnectiontype['Calculation Formula']='Fix value' 
    #                     newconnectiontype['Performance']=herraje['Performance']  
    #                     newconnectiontype['is_modeled']='modeled' 
    #                     recuentodeconnectiontypes.append(herraje['connection_type'])
    #                     finalconnectiontypes.append(newconnectiontype)
    #             if herraje['parentjoint_id']==connectiontype['parentjoint_id'] and herraje['connection_type']!=connectiontype['connection_type']:
    #             #si coinciden el parent y pero el connectiontype es diferente    
    #                 if herraje['connection_type'] not in recuentodeconnectiontypes:
    #                     newconnectiontype={}
    #                     newconnectiontype['Calculation Formula']='Fix value' 
    #                     newconnectiontype['Performance']=herraje['Performance'] 
    #                     newconnectiontype['connection_type']=herraje['connection_type']
    #                     newconnectiontype['connectiongroup_type_id']='' 
    #                     newconnectiontype['is_modeled']='modeled' 
    #                     newconnectiontype['parentjoint_id']=herraje['parentjoint_id']
    #                     recuentodeconnectiontypes.append(herraje['connection_type'])
    #                     finalconnectiontypes.append(newconnectiontype)
                
                    
    #     if connectiontype['is_modeled'][0]=='No':
    #         finalconnectiontypes.append(connectiontype)
                        
    # connectiongroup_cost = 0
    # for connectiontype in finalconnectiontypes:
    #     connectiontype_id=connectiontype['connection_type']
    #     connectiontype_calcform=connectiontype['Calculation Formula']
    #     connectiontype_performance=connectiontype['Performance']          
    #     allclayers=airtable_clayers
    #     filteredclayers=[]
        
    #     for i in allclayers:
    #         if connectiontype_id in i['connection_type_code']:
    #             i['parentjoint_id']=parentid
    #             filteredclayers.append(i)
                
    #     for clayer in filteredclayers:
    #         material_performance=clayer['Performance']
    #         material_formula=clayer['Calculation Formula']
    #         material_sku=clayer['material_id']
    #         material_cost=clayer.get('Current_material_cost',[0])[0]
    #         material_unit=clayer['Units'][0]                     
            
    #         if material_formula=='Fix value':
    #             if material_unit=='U':
    #                 if connectiontype_calcform=='Length * performance':
    #                     quantity=math.ceil(material_performance*connectiontype_performance*long)
    #                     connectiongroup_cost=connectiongroup_cost+material_cost*quantity
    #                     clayer['quantity']=quantity
    #                     clayer['layer_cost']=material_cost*quantity
    #                 elif connectiontype_calcform=='Opening * performance':
    #                     quantity=material_performance*connectiontype_performance*openings
    #                     connectiongroup_cost=connectiongroup_cost+material_cost*quantity                        
    #                     clayer['quantity']=quantity
    #                     clayer['layer_cost']=material_cost*quantity                        
    #                 else:
    #                     quantity=material_performance*connectiontype_performance
    #                     connectiongroup_cost=connectiongroup_cost+material_cost*quantity                        
    #                     clayer['quantity']=quantity
    #                     clayer['layer_cost']=material_cost*quantity
    #             else:
    #                 if connectiontype_calcform=='Length * performance':
    #                     quantity=material_performance*connectiontype_performance*long
    #                     connectiongroup_cost=connectiongroup_cost+material_cost*quantity                        
    #                     clayer['quantity']=quantity
    #                     clayer['layer_cost']=material_cost*quantity
    #                 elif connectiontype_calcform=='Opening * performance':
    #                     quantity=material_performance*connectiontype_performance*openings
    #                     connectiongroup_cost=connectiongroup_cost+material_cost*quantity                        
    #                     clayer['quantity']=quantity
    #                     clayer['layer_cost']=material_cost*quantity
    #                 else:
    #                     quantity=material_performance*connectiontype_performance
    #                     connectiongroup_cost=connectiongroup_cost+material_cost*quantity                        
    #                     clayer['quantity']=quantity
    #                     clayer['layer_cost']=material_cost*quantity
    #         inferredandmodeledclayers.append(clayer)        
            
    # return connectiongroup_cost, inferredandmodeledclayers        
        
        
        
    
        
        
        